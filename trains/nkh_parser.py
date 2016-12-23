import os
import pickle
import gzip

from utility import roots

ABC = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
TMPFL = "nkhtmp.pkl.gz"
TMPROOT = roots["tmp"]
HEADERFL = "nkh_header.csv"


def valid(cell):
    return cell.value not in (None, "", " ")


def parse_wb(workbook):
    def extract_data(worksheet):

        def header_infer(ws):
            # print("... infering header location ...")
            topleft = ""
            start = str()
            end = str()
            record = []

            for r in ws.iter_rows():
                for i, cll in enumerate(r):
                    if cll.value not in (None, "", " "):
                        if not topleft:
                            topleft = cll.value, cll.coordinate
                            continue
                        record.append(str(cll.value))
                        if not start:
                            start = cll.coordinate
                            break
                if start:
                    break

            # headrow = tuple(ws.rows[int(start[-1]) - 1])
            headrow = [row for i, row in enumerate(ws.iter_rows()) if i == int(start[-1]) - 1][0]
            cll = None
            if all(map(lambda c: valid(c), headrow)):
                end = headrow[-1].coordinate
            else:
                for i, cll in enumerate(headrow):
                    if not valid(cll):
                        if all(map(lambda c: not valid(c), headrow[i:])):
                            end = headrow[i - 1].coordinate
                            break
            if not end and cll is not None:
                end = cll.coordinate

            return record, start, end

        def data_infer(ws, hstart):
            # print("... infering data location ...")
            end = ""
            # col = tuple(ws.columns[ABC.index(hstart[0])])
            col = [c for i, c in enumerate(ws.columns) if i == ABC.index(hstart[0])][0]
            if all(map(lambda cell: valid(cell), col)):
                return col[-1].coordinate

            cll = None
            startrow = int(hstart[1]) - 1
            for i, cll in enumerate(col):
                if i < startrow:
                    continue
                if not cll.value:
                    if all(map(lambda cell: not valid(cell), col[i:])):
                        end = col[i - 1].coordinate
                        break
            if not end and cll is not None:
                end = cll.coordinate
            assert end is not None, "OMG it's all wrong!"
            return end

        def data_extract(ws, rng: str):
            # print("... extracting data ...")
            matrix = []
            for row in ws[rng]:
                datarow = []
                for cell in row:
                    datarow.append(str(cell.value))
                matrix.append(datarow)
            return matrix

        header, inferred_header_start, inferred_header_end = header_infer(worksheet)
        inferred_data_end = data_infer(worksheet, inferred_header_start)
        inferred_range = "{}:{}".format(inferred_header_start, inferred_header_end[0] + inferred_data_end[1:])

        # print("DATA RANGE:\n  QUERIED: {}\n  INFERRED: {}".format(worksheet.dimensions, inferred_range))
        data = data_extract(worksheet, inferred_range)

        return data

    objecttype_data = dict()
    for sheetname in workbook.get_sheet_names():
        if sheetname.lower() == "01" or "forrásadat" in sheetname.lower() or "saját célú" in sheetname.lower():
            continue
        print("EXTRACTING FROM WS: {}".format(sheetname))
        objecttype_data[sheetname.lower()] = extract_data(workbook[sheetname])

    return objecttype_data


def merge_data_by_objecttype(data: dict):
    outchains = dict()
    for line_code, dictionary in data.items():
        for objecttype, matrix in dictionary.items():
            if objecttype not in outchains:
                outchains[objecttype] = "Vonalszám\t"
                for i, row in enumerate(matrix):
                    if i > 0:
                        outchains[objecttype] += str(line_code)[:6] + "\t"
                    for value in row:
                        outchains[objecttype] += value + "\t"
                    outchains[objecttype] += "\n"
            else:
                for row in matrix[1:]:
                    outchains[objecttype] += str(line_code)[:6] + "\t"
                    for value in row:
                        outchains[objecttype] += value + "\t"
                    outchains[objecttype] += "\n"
    return outchains


def assert_headers_are_the_same(data: dict):
    def dump_headers():
        outchain = ""
        for ot, hd in obj_header.items():
            outchain += ot + "\t" + "\t".join(hd) + "\n"
        with open(roots["headers"] + HEADERFL, "w") as headfl:
            headfl.write(outchain.replace("None", ""))
            headfl.close()

    output = True
    obj_header = dict()
    for line_number, dictionary in data.items():
        for objecttype, matrix in dictionary.items():
            if objecttype not in obj_header:
                obj_header[objecttype] = matrix[0]
            else:
                output = all([ot1 == ot2] for ot1, ot2 in
                             zip(obj_header[objecttype], matrix[0]))
                if not output:
                    print("Deviance at VSz: {}, OT: {}".format(line_number, objecttype))
    assert output, "Deviances were detected"
    print("Headers asserted! They're all the same...")
    if HEADERFL not in os.listdir(roots["headers"]):
        dump_headers()
        print("Headers written to {}".format(roots["headers"] + HEADERFL))


def dump_data(outchains_dictionary: dict, outroot: str):
    if outroot[-1] not in ("/", "\\"):
        outroot += "/"
    for objecttype, outchain in outchains_dictionary.items():
        flname = str(objecttype) + ".csv"
        print("Writing {}".format(flname))
        with open(outroot + flname, "w") as outfl:
            outfl.write(outchain.replace("None", ""))
            outfl.close()


def pull_from_xls():
    import openpyxl as xl
    os.chdir(roots["nkh"] + "Halmai Zsolttól/NKH/Feltöltött/")

    files = [flname for flname in os.listdir(".") if ".xls" in flname]
    objecttypes = set()
    linenumber_objecttype_names = dict()

    for flname in sorted(files):
        print("PARSING:", flname)
        wb = xl.load_workbook(flname, use_iterators=True)
        sheetnames = set([shname.lower() for shname in wb.get_sheet_names()])
        assert len(sheetnames) == len(wb.get_sheet_names()),\
            "Object-type listed more than once in {}".format(flname)
        if not objecttypes:
            objecttypes = sheetnames
        else:
            # assert sheetnames == objecttypes,\
            #     "Different object types from previous ones in {}!".format(flname)
            pass
        linenumber_objecttype_names[flname] = parse_wb(workbook=wb)

    return linenumber_objecttype_names


if __name__ == '__main__':
    if TMPFL not in os.listdir(TMPROOT):
        print("Cached file not found... Rereading...")
        nested = pull_from_xls()
        print("Dumping cache file...")
        with gzip.open(TMPROOT + TMPFL, mode="wb") as fl:
            pickle.dump(nested, fl)
            fl.close()
    else:
        print("Cache file found. Reading data...")
        with gzip.open(TMPROOT + TMPFL, "rb") as fl:
            nested = pickle.load(fl)
            fl.close()

    assert_headers_are_the_same(nested)
    objecttype_merged = merge_data_by_objecttype(nested)
    dump_data(outchains_dictionary=objecttype_merged, outroot=roots["data"] + "Output/")
    print("Finite Incantatum!")
    print("The following statements were asserted for being true:")
    print("- Whether every NKH workbook contains every objecttype-titled worksheet.")
    print("- Whether an objecttype is listed more than once in a workbook.")
    print("- Whether the title of a worksheet (the objecttype name) matches the value in [A1]")
    print("- Whether the ending column's number of the data is the same inferred\n  by openpyxl and by myself.")
    print("- Whether every header for a certain objecttype is the same")
