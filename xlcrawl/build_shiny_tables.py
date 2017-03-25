import os
from collections import defaultdict

import numpy as np

import openpyxl as xl
from openpyxl import Workbook
from openpyxl.worksheet import Worksheet

from SciProjects.xlcrawl import project_root, shinyroot

templates = {c: shinyroot + c + "_template.xlsx" for c in ("chem", "item", "inst")}
categories = ("chem", "item", "inst")


class Reorder:

    @staticmethod
    def tonice(line):
        return line[:-1].split("\t")

    @staticmethod
    def chem(line):
        """
        Source head: FILE; Vegyszer neve; CAS; Tisztaság; Szüks. menny; Cikkszám
        Target head: Cikkszám; Vegyszer; Mennyiségi egység; Felhasznált mennyiség
        """
        return [line[5], line[1], "", line[4]]

    @staticmethod
    def inst(line):
        """
        Source head: FILE; OLDFIELD; NEWFIELD; USAGE; INVNUMBERS
        Target head: Laboratóriumi műszer; Leltári szám; Gyári szám; Vizsgálatok száma
        """
        return [line[2], "; ".join(d for d in line[4:] if d), "", line[3]]

    @staticmethod
    def item(line):
        """
        Source head: FILE; OLDFIELD; NEWFIELD; QUANT; INVNUMBERS
        Target head: Cikkszám; Szakmai készlet; db/vizsgálatok száma
        """
        return ["; ".join(d for d in line[4:] if d), line[2], line[3]]

    @staticmethod
    def do(line, categ):
        return {"chem": Reorder.chem, "inst": Reorder.inst,
                "item": Reorder.item}[categ](line)


def pull_and_reorder_source(categ):
    data = defaultdict(list)
    with open(shinyroot + categ + "_parsed.csv") as handle:
        next(handle)  # skip header row
        for line in handle:
            line = Reorder.tonice(line)
            data[line[0]].append(Reorder.do(line, categ))
        data = {k: np.array(v) for k, v in data.items()}
    return data


chemdata, itemdata, instdata = tuple(map(pull_and_reorder_source, categories))


def instantiate_template(flnm, categ):
    newpath = shinyroot + "reparsed_" + categ + "/" + flnm
    handle = xl.load_workbook(templates[categ])
    handle.save(newpath)
    return xl.load_workbook(newpath)


def dump_to_xlfl(handle: Workbook, flnm, categ):
    if categ == "chem":
        source = chemdata
    elif categ == "inst":
        source = instdata
    else:
        source = itemdata
    ws = handle.get_sheet_by_name("Adat")  # type: Worksheet
    try:
        matrix = source[flnm]
    except KeyError:
        print("Couldn't get [{}] for {}".format(categ, flnm))
        matrix = np.array([["-"]*(3 + int(categ != "item"))])
    Xn, Yn = matrix.shape

    offset = int(categ == "inst")
    cellz = np.array(list(ws.iter_rows(min_row=6, max_row=5+Xn, min_col=offset+1, max_col=offset+Yn)))
    assert cellz.shape == matrix.shape, "c: {} m: {}".format(cellz.shape, matrix.shape)
    for c, d in zip(cellz.ravel(), matrix.ravel()):
        c.value = d
    handle.save(shinyroot + "reparsed_" + categ + "/" + flnm)


if __name__ == '__main__':
    for filename in os.listdir(project_root + "ALLXLFLZ/"):
        for category in ("chem", "item", "inst"):
            print("Doing [{}] of {}".format(category, filename))
            handle = instantiate_template(filename, category)
            dump_to_xlfl(handle, filename, category)
