import os
import pickle as pkl

import openpyxl as xl

from SciProjects.xlcrawl import projectroot
from openpyxl.worksheet import Worksheet


mapping = pkl.load(open(projectroot + "instusage.pkl", "rb"))


def valid_iid(string):
    return str(string).replace(" ", "")


def walktable(ws: Worksheet, flnm):
    for rown in range(6, 50):
        iid = valid_iid(ws[f"A{rown}"].value)
        if iid == "None":
            continue
        try:
            ws[f"D{rown}"].value = mapping[iid]
        except KeyError:
            print("Key error @", flnm)
            return


def main():
    for xlfl in (fl for fl in os.listdir(projectroot + "FIN/") if fl[-5:] == ".xlsx"):
        wb = xl.load_workbook(projectroot + "FIN/" + xlfl, data_only=True)
        walktable(wb["inst"], xlfl)
        wb.save(projectroot + xlfl)
    print("-- END PROGRAM --")

if __name__ == '__main__':
    main()
