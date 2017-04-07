import os
from difflib import SequenceMatcher

import openpyxl as xl
from openpyxl.worksheet import Worksheet

from SciProjects.xlcrawl import projectroot


def strsim(str1, str2):
    return SequenceMatcher(None, str1, str2).ratio()


def striters(*iterables):
    out = []
    for iterable in iterables:
        itt = type(iterable)
        out.append(itt(map(str, map(lambda d: "-" if d is None else d, iterable))))
    return out if len(out) > 1 else out[0]


def valid(ws: Worksheet):
    return "melléklet" in "".join((str(ws["A1"].value), str(ws["A2"].value))).lower()


def iter_flz(dirroot):
    os.chdir(dirroot)
    paths = []
    print("Gathering information...", end=" ")
    for root, dirz, flz in os.walk("."):
        paths += [os.path.join(root, fl) for fl in flz if fl[-4:] == "xlsx" and fl[0] != "~"]
    print("Done!")
    print("Reading XL files...")
    N = len(paths)
    strlen = len(str(N))
    for i, xlfl in enumerate(sorted(paths), start=1):
        print("{:>{w}}/{} - {}".format(i, N, xlfl, w=strlen))
        os.chdir(dirroot)
        ws = xl.load_workbook(xlfl).worksheets[0]
        if not valid(ws):
            print(" !!! iter_flz: skipping invalid file:", xlfl)
            continue
        yield ws, xlfl


def walk_column_until(ws: Worksheet, coln: int, strval: str, limit=30, limitstr=""):
    strval = strval.lower()
    itr = next(ws.iter_cols(min_col=coln+1, max_col=coln+1, min_row=0, max_row=limit))
    for i, c in enumerate(itr):
        cval = str(c.value).lower()
        if strval in cval:
            return i+1
        if i >= limit:
            return None
        if limitstr:
            if limitstr in cval:
                return None


def extract_inventory_numbers(raw_field: str):
    if len(raw_field) < 6:
        return [], raw_field
    newfield = (raw_field
                .replace("NAVG ", "NAVG")
                .replace("L ", "L")
                .replace("V ", "V")
                .replace("VV ", "V")
                .replace("LL ", "L")
                .replace("(", " ")
                .replace(")", " ")
                .replace("[", " ")
                .replace("]", " ")
                .replace("•", "")
                .replace("*", "")
                .replace("+", "")
                .replace("·", "")
                .replace("/", " "))
    if newfield[0] == "-":
        newfield = newfield[1:]
    newfield = trimstr(newfield)
    words = newfield.split(" ")
    candidates = []
    for w in words:
        w = trimstr(w).replace(" ", "").replace(",", "").replace(".", "").upper()
        if len(w) < 6:
            continue
        if w[0] in "LVlv" and w[1:].isdigit():
            candidates.append(w)
        elif w[:4] == "NAVG":
            w.replace(" ", "")
            candidates.append(w)
        elif w.isdigit():
            candidates.append(w)
    for candidate in candidates:
        newfield = newfield.replace(candidate, "").replace(" ,", ",").replace(" ;", ";")
    while "  " in newfield:
        newfield = newfield.replace("  ", " ")
    while ",," in newfield:
        newfield = newfield.replace(",,", ",")
    while ";;" in newfield:
        newfield = newfield.replace(";;", ",")
    return candidates, newfield


def trimstr(ugly):
    return (ugly
            .replace("(", "")
            .replace(")", "")
            .replace("/", "")
            .replace(":", "")
            .replace(";", "")
            .replace("  ", " ")
            .strip())


class DJ:

    def __init__(self, xlpath=projectroot+"Díjjegyzék.xlsx"):
        self.djnames = {}
        self.munumbers = {}
        self.munames = {}
        djwb = xl.load_workbook(xlpath)
        for row in djwb.worksheets[0]:
            djnum = row[1].value
            if not isinstance(djnum, int):
                continue
            self.djnames[djnum] = row[2].value
            self.munumbers[djnum] = row[5].value
        # for row in djwb.worksheets[1]:
        #     self.munames[row[0].value] = row[1].value

    def dj_name_to_nums(self, name):
        return [k for k, v in sorted(self.djnames.items()) if name == v]

    def mu_to_dj(self, mu):
        return [k for k, v in sorted(self.munumbers.items()) if v == mu]


class Allomany:

    def __init__(self, src=None):
        handle = open(src if src is not None else projectroot + "allomany.csv")
        dstream = (d.strip().split("\t")[:3] for d in handle)
        self.data = {nev: (tasz, helyettes) for nev, tasz, helyettes in dstream}

    def get(self, what, nev):
        result = self.data.get(nev, ("", ""))
        return result[int(what=="hely")]

    def tasz(self, nev):
        return self.get("tasz", nev)

    def helyettes(self, nev):
        return self.get("hely", nev)
