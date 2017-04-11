import openpyxl as xl

from SciProjects.xlcrawl import projectroot
from SciProjects.xlcrawl.assemble_database import valid_iid


chain = """Key error @ 236-Olajmagvak olajtartalmának meghatározása.xlsx
Key error @ 163-Szárazanyag-tartalom meghatározása szárítószekrényben.xlsx
Key error @ 216-Szterintartalom meghatározás.xlsx
Key error @ 159-Polifenol oxidáz aktivitás kimutatása (hőkezeltség vizsgálat).xlsx
Key error @ 147-Szárazanyag-tartalom meghatározása szárítószekrényben.xlsx
Key error @ 144-Sajtok, tejtermékek és túró szárazanyag-tartalmának meghatározása.xlsx
Key error @ 146-Vaj zsírtartalmának meghatározása.xlsx
Key error @ 145-Sajtok, tejtermékek és túró zsírtartalmának meghatározása.xlsx
Key error @ 174-Tejfehérje mennyiségi meghatározása elektroforézissel.xlsx
Key error @ 205-Peroxidáz aktivitás kimutatása (hőkezeltség vizsgálat).xlsx
Key error @ 107-Borok alkoholtartalmának meghatározása.xlsx
Key error @ 157-Nyersrost-tartalom meghatározás.xlsx
Key error @ 170-Kekszek, édesipari termékek szárazanyag-tartalmának meghatározása.xlsx
Key error @ 242-Zsírtartalom meghatározása Soxhlet extrakcióval.xlsx
Key error @ 155-Fehérje izolátumok, lisztek -szója, rizs, egyéb- azonosító vizsgálata a fehérjék molekulatömeg szerinti elválasztása alapján – elektroforézis alkalmazásával.xlsx
Key error @ 223-Zsírtartalom meghatározása Soxhlet extrakcióval.xlsx
Key error @ 173-Tejfehérje meghatározása tejcsokoládé mintákból.xlsx
Key error @ 187-Fehérje izolátumok, lisztek -szója, rizs, egyéb- azonosító vizsgálata a fehérjék molekulatömeg szerinti elválasztása alapján -elektroforézis-alkalmazásával.xlsx
Key error @ 158-Peroxidáz aktivitás kimutatása (hőkezeltség vizsgálat).xlsx
Key error @ 259-Takarmányok nyersolaj és zsírtartalmának meghatározása.xlsx
Key error @ 201-Nyersrost-tartalom meghatározás.xlsx
Key error @ 257-Keményítőtartalom meghatározása enzimatikus módszerrel állati takarmányokból.xlsx
Key error @ 202-Oldhatósági vizsgálat.xlsx
Key error @ 238-Vaj zsírtartalmának meghatározása.xlsx
Key error @ 88-Alkoholtartalom meghatározása desztillációval és elektronikus sűrűségméréssel.xlsx
Key error @ 258-Keményítőtartalom meghatározása. Ewers-féle polarimetriás módszer.xlsx
Key error @ 103-Összes cián-hidrogéntartalom meghatározása.xlsx
Key error @ 169-Kekszek, édesipari termékek zsírtartalmának meghatározása.xlsx
Key error @ 217-Tejfehérje mennyiségi meghatározása elektroforézissel.xlsx
Key error @ 222-Zsírtartalom meghatározása Schmid-Bordzynski-Ratzlaff (SBR) módszerrel.xlsx
Key error @ 227-BLG mennyiségi meghatározása ELISA technikával.xlsx
Key error @ 138-Szárazanyag-tartalom meghatározása szárítószekrényben.xlsx
Key error @ 233-Marhahús genomiális vizsgálata.xlsx
Key error @ 241-Zsírtartalom meghatározása Schmid-Bordzynski-Ratzlaff (SBR) módszerrel.xlsx
Key error @ 194-Keményítőtartalom meghatározása enzimatikus módszerrel.xlsx
Key error @ 203-Összes cián-hidrogéntartalom meghatározása.xlsx
Key error @ 143-Juhtejből, kecsketejből, vagy bivalytejből, illetve juh-, kecske- vagy bivalytej keverékéből készített sajtokban található tehéntej és kazeinát kimutatása.xlsx
Key error @ 237-Takarmányok nyersolaj és zsírtartalmának meghatározása.xlsx
Key error @ 196-Keményítőtartalom meghatározása. Ewers-féle polarimetriás módszer.xlsx
Key error @ 235-Nyersolajok és zsírok meghatározása.xlsx
Key error @ 154-Szárazanyag tartalom meghatározása vákuum-szárítószekrényes módszer.xlsx
Key error @ 248-Keményítőtartalom meghatározása. Ewers-féle polarimetriás módszer.xlsx
Key error @ 168-Kazein mennyiségi meghatározása ELISA technikával.xlsx
Key error @ 226-Baromfihús tartalom kimutatása hústermékekből, fajspecifikus ELISA technika alkalmazásával.xlsx
Key error @ 249-Oldhatósági vizsgálat.xlsx
Key error @ 246-Keményítőtartalom meghatározása enzimatikus módszerrel.xlsx
Key error @ 192-Kazein mennyiségi meghatározása ELISA technikával.xlsx
Key error @ 206-Polifenol oxidáz aktivitás kimutatása (hőkezeltség vizsgálat).xlsx
Key error @ 214-Szárazanyag tartalom meghatározása vákuum-szárítószekrényes módszer.xlsx
Key error @ 171-Keményítőtartalom meghatározása enzimatikus módszerrel.xlsx""".split("\n")

xlflz = [l.split(" @ ")[-1] for l in chain]

for xlfl in xlflz:
    print(xlfl)
    wb = xl.load_workbook(projectroot + "FIN/" + xlfl, data_only=True)
    ws = wb["inst"]
    for rown in range(6, 50):
        row = [ws.cell(row=rown, column=coln).value for coln in range(1, 10)]
        if all(c is None for c in row):
            continue
        if not valid_iid(row[0]) and row[0] is not None:
            if valid_iid(row[1]):
                print("SWAPPED")
                ws[f"A{rown}"] = row[1]
                ws[f"B{rown}"] = row[0]
        if "None" in row:
            print("DELETING NONES")
            for i, c in enumerate(row, start=1):
                if c == "None":
                    ws.cell(row=rown, column=i).value = None
    wb.save(projectroot + "FIN/" + xlfl)
