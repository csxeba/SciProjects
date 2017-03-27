import os

from SciProjects.xlcrawl import projectroot
from SciProjects.xlcrawl.util import walk_column_until, iter_flz
from openpyxl.worksheet import Worksheet


def locate_table(ws: Worksheet, flnm):
    myrow = walk_column_until(ws, 0, "szükséges műszer", limit=120)
    tablerow = None
    if myrow is None:
        print("No <szükséges műszer> header in {}".format(flnm))
        return None
    coln = 0
    while tablerow is None:
        if coln >= 20:
            return None
        cells = next(ws.iter_cols(min_col=coln, max_col=coln, min_row=myrow, max_row=myrow+10))
        for i, c in enumerate(cells, start=myrow):
            if "ráosztott költség" in str(c.value) or i >= myrow+10:
                break
            if c.value is not None:
                tablerow = i + 1
                break
        coln += 1
    return tablerow, coln


def extract_row(ws, rown, startcol):
    data = []
    coloffset = 0
    while coloffset != 20:
        found = ws.cell(row=rown, column=startcol + coloffset).value
        data.append(str(found).strip() if found is not None else None)
        coloffset += 1
    return list(filter(lambda x: x is not None, data))


def extract_matrix(ws: Worksheet, row, col, flnm):
    rown = row
    data = []
    ran = 0
    limit = 50
    first_empty_row = 0
    while ran < limit:
        if ran >= limit:
            if row:
                limit += 1
        col0 = str(ws.cell(row=rown, column=1).value)
        row = extract_row(ws, rown, col)
        if "ráosztott költség" in col0:
            break
        if not row:
            if not first_empty_row and ran > 1:
                first_empty_row = len(data)
            ran += 1
            rown += 1
            continue
        data.append(row)
        ran += 1
        rown += 1
    if len(data) == 0:
        data = [[""]*2]
    return data


def main():
    os.chdir(project_root + "/ALLXLFLZ/")

    lndir = len(os.listdir("."))
    strln = len(str(lndir))
    chain = "FILE\tINSTRUMENT\tUSAGE\tOTHER_INFO->\n"
    for i, (flnm, ws) in enumerate(iter_flz("."), start=1):
        print("\rProgress: {:>{w}}/{} @ {}".format(i, lndir, flnm, w=strln), end="")
        row, col = locate_table(ws, flnm)
        data = extract_matrix(ws, row, col, flnm)
        if not data:
            chain += flnm + "\t" + "\t".join("-" for _ in range(3)) + "\n"
            continue
        chain += "\n".join("\t".join(line) for line in data) + "\n"

    with open(project_root + "instrument.csv", "w") as handle:
        handle.write(chain.replace("None", "-"))

if __name__ == '__main__':
    main()
