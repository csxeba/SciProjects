import numpy as np
import openpyxl as xl
from openpyxl.worksheet import Worksheet

from SciProjects.xlcrawl import projectroot, templateroot
from SciProjects.xlcrawl.util import DJ, Allomany

root = projectroot + "BobeTables/"

dj = DJ()
personell = Allomany()

def pullinfo(flpath):
    ws = xl.load_workbook(flpath).worksheets[0]
    data = []
    for row in ws.rows:
        line = []
        for cell in row:
            line.append(str(cell.value).strip())
        if not line:
            continue
        data.append(line)
    data = np.array(data)
    djnumz = set(data[1:, 0])
    return data[1:], djnumz


def reparse_names(data, djnum):
    for line in data:
        if line[0] == str(djnum):
            # djn, mernok, mtasz, mtime, techni, ttasz, ttime
            return line


def reparse_tab(data, djnum):
    return [l[1:] for l in data if l[0] == str(djnum)]


def dump_header_data(djn, namereparsed, xlwb):
    djname = dj.djnames.get(int(djn), "")
    mnum = dj.munumbers.get(int(djn), "")
    mname, tname = namereparsed[1], namereparsed[4]
    mtasz, ttasz = personell.tasz(mname), personell.tasz(tname)
    ws = xlwb["head"]
    ws["B4"].value = djn
    ws["B5"].value = djname
    ws["B6"].value = mname
    ws["B9"].value = djname
    ws["B11"].value = mnum
    ws["B14"].value = mname
    ws["B19"].value = tname


def dump_table(tab, name, xlwb: xl.Workbook):
    ws = xlwb.get_sheet_by_name(name)  # type: Worksheet
    for i, line in enumerate(tab, start=6):
        for j, data in enumerate(line, start=1):
            ws.cell(row=i, column=j).value = data


def dump(djn, headdata, chemdata, instdata, itemdata):
    template = xl.load_workbook(templateroot + "template.xlsx")
    template.save(root + "FIN/" + djn + ".xlsx")
    del template
    outwb = xl.load_workbook(root + "FIN/" + djn + ".xlsx")
    dump_header_data(djn, reparse_names(headdata, djn), outwb)
    dump_table(reparse_tab(chemdata, djn), "chem", outwb)
    dump_table(reparse_tab(itemdata, djn), "item", outwb)
    dump_table(reparse_tab(instdata, djn), "inst", outwb)
    outwb.save(root + "FIN/" + djn + ".xlsx")


def main():
    pulled = [pullinfo(root + path + ".xlsx") for path in ("chem", "inst", "item")]
    tabz, djnz = list(zip(*pulled))
    hours, hdjn = pullinfo(root + "hrs.xlsx")
    for djn in sorted(list(hdjn)):
        print("\rDoing DJN: {:>4}".format(djn), end="")
        dump(djn, hours, *tabz)
    print()

if __name__ == '__main__':
    main()
