import os
import pickle

import numpy as np
from scipy.stats import mode

import openpyxl as xl
from openpyxl.worksheet import Worksheet

from SciProjects.xlcrawl import projectroot
from SciProjects.xlcrawl.util import headranges

finalzroot = projectroot + "FIN/"

xlstream = (xlfl for xlfl in os.listdir(finalzroot) if xlfl[-5:] == ".xlsx")


def getID(wb: xl.Workbook, flnm):
    ws = wb["head"]
    djn = str(ws[headranges["djn"]].value)
    if "." in djn:
        djn = djn.replace(".", "")
        if not djn.isdigit():
            raise RuntimeError("DJN: " + djn + " problems with " + flnm)
        ws[headranges["djn"]].value = int(djn)
        wb.save(finalzroot + flnm)
    if not djn.isdigit():
        djn = flnm.split("-")[0]
        assert djn[0] == "U"
    return str(djn)


def valid_iid(string):
    string = str(string)
    return any((string[0].isdigit(), string[0] == "L", string[:3] == "NAV"))


def parse_ws(ws: Worksheet, flnm):
    nilz = 0
    data = []
    for rown in range(6, 50):
        row = [ws.cell(row=rown, column=coln).value for coln in range(1, 5)]
        assert len(row) == 4
        if all(c is None for c in row):
            nilz += 1
            if nilz >= 10:
                # print("Empty sheet @", flnm)
                break
            continue
        usage = str(row[-1]).split(" ")[0].split("/")[0].strip()
        if not usage.isdigit():
            usage = 0
        row[-1] = usage
        row[0] = row[0] if row[0] is None else str(row[0]).strip()
        if not valid_iid(row[0]) and row[0] is not None:
            if valid_iid(row[1]):
                # print("Swapped @", flnm)
                iid = row[1]
                row[1] = row[0]
                row[0] = iid
            else:
                print("Uncorrectabe: {} @ {}".format(row[0], flnm))
        data.append(list(map(str, row)))
    return data


def dumpinstusage(data: np.array):
    db = {}
    for iid in np.unique(data[:, 0]):
        if iid == "None" or not iid:
            continue
        args = np.argwhere(data[:, 0] == iid)
        d = data[args, -1].astype(float).ravel()
        d[d == 0.] = np.nan
        db[iid] = int(mode(d, nan_policy="omit")[0][0])
        # db[iid].append(mode(d, nan_policy="omit")[0][0])/
    outchain = "\n".join(f"{k}\t{db[k]}" for k in sorted(db))
    with open(projectroot + "instusage.csv", "w") as csvhandle:
        csvhandle.write(outchain)
    with open(projectroot + "instusage.pkl", "wb") as pklhandle:
        pickle.dump(db, pklhandle)


def getdata():
    data = []
    allxl = len(os.listdir(finalzroot))
    strlen = len(str(allxl))
    for i, xlfl in enumerate(xlstream, start=1):
        print("\r{:>{w}}/{}".format(i, allxl, w=strlen), end="")
        wb = xl.load_workbook(finalzroot + xlfl, data_only=True)
        table = parse_ws(wb["inst"], xlfl)
        if table:
            data.extend(table)
    print("Dumping numpy array...")
    data = np.array(data)
    np.save(projectroot + "datasaved.npy", data)
    return data


def main():
    if os.path.exists(projectroot + "datasaved.npy"):
        data = np.load(projectroot + "datasaved.npy")
    else:
        data = getdata()
    assert data.ndim == 2
    print("Dumping to csv...")
    with open(projectroot + "instusage_full.csv", "w") as handle:
        handle.write("\n".join("\t".join(line) for line in data))
    print("Assembling...")
    dumpinstusage(data)
    print(" -- END PROGRAM -- ")

if __name__ == '__main__':
    main()
