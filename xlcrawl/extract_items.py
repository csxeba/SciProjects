import os

from SciProjects.xlcrawl import projectroot
from SciProjects.xlcrawl.util import walk_column_until, iter_flz
from openpyxl.worksheet import Worksheet


def locate_table(ws: Worksheet, flnm):
    myrow = walk_column_until(ws, 0, "eszközök, műszerek", limit=50)
    tablerow = None
    if myrow is None:
        print("No <eszköz, műszer> header in {}".format(flnm))
        return None
    coln = 0
    while tablerow is None:
        if coln >= 20:
            return None
        cells = next(ws.iter_cols(min_col=coln, max_col=coln, min_row=myrow, max_row=myrow+10))
        for i, c in enumerate(cells, start=myrow):
            if "szükséges műszer" in str(c.value) or i >= myrow+10:
                break
            if c.value is not None:
                tablerow = i + 1
                break
        coln += 1
    return tablerow, coln


def extract_row(ws, rown, startcol):
    data = []
    coloffset = 0
    while coloffset != 20:
        found = ws.cell(row=rown, column=startcol + coloffset).value
        data.append(str(found).strip() if found is not None else None)
        coloffset += 1
    return list(filter(lambda x: x is not None, data))


def extract_matrix(ws: Worksheet, row, col, flnm):
    rown = row
    data = []
    ran = 0
    first_empty_row = 0
    lines_found = 0
    while 1:
        if ran > 50:
            print("Overrun in {}.\nClipping to {}!".format(flnm, first_empty_row))
            return data[:first_empty_row]
        col0 = str(ws.cell(row=rown, column=1).value)
        row = extract_row(ws, rown, col)
        if "szükséges műszer" in col0:
            break
        if not row:
            if rown > 1 and not first_empty_row:
                first_empty_row = lines_found
            ran += 1
            rown += 1
            continue
        data.append(row)
        lines_found += 1
        ran += 1
        rown += 1
    if len(data) == 0:
        data = [[""]*2]
    return data


def assemble_item_data(output=None):
    os.chdir(projectroot + "ALLXLFLZ")
    lndir = len(os.listdir("."))
    strln = len(str(lndir))
    chain = "FILE\tNAME\tQUANT\tOTHER_INFO->"
    for i, (flnm, ws) in enumerate(iter_flz("."), start=1):
        print("\rProgress: {:>{w}}/{} @ {}".format(i, lndir, flnm, w=strln), end="")
        row, col = locate_table(ws, flnm)
        data = extract_matrix(ws, row, col, flnm)
        if not data:
            chain += flnm + "\t" + "\t".join("-" for _ in range(3)) + "\n"
            continue
        chain += "\n".join("\t".join(line) for line in data) + "\n"

    with open(output, "w") as handle:
        handle.write(chain.replace("None", "-"))

if __name__ == '__main__':
    assemble_item_data()
