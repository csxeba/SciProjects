import os
import gzip
import pickle

import numpy as np
import openpyxl as xl
import time

from SciProjects.xlcrawl import projectroot, templateroot, pickleroot, sourceroot
from SciProjects.xlcrawl.util import extract_inventory_numbers, iter_flz
from openpyxl.worksheet import Worksheet

# vbhvcfrcd  bh jn -- Ezt a Cuni írta :)


def _slicepath(flpath):
    return os.path.split(flpath)


class NamelessAbstraction:

    templates = {
        c: templateroot + c + "_template.xlsx"
        for c in ("chem", "item", "inst", "head")
        }

    def __init__(self, path, flnm, categ, data):
        self.path = path
        self.flnm = flnm
        self.categ = categ
        self.data = data
        self.reparsed = None


class Table(NamelessAbstraction):

    """
    Raw extracted data is shaped like this:
    CHEM: NAME; CAS; PURITY; QUANT; INVNUM
    ITEM: NAMESTRING (INVNUM); QUANT
    INST: NAMESTRING (INVNUM); QUANT?

    Reparsed data is shaped like this:
    CHEM: INVNUM; NAME; QUANT_UNIT; QUANT_USED
    ITEM: INVNUM; NAME; QUANT_USED
    INST: INVNUM; NAME; SN; QUANT_USED
    """

    def __init__(self, data, path, flnm, categ):
        super().__init__(path, flnm, categ, data)
        self.dim = 3 if categ == "item" else 4
        if isinstance(data, Worksheet):
            self.read_data(data)
            self.reparse_data()

    def empty(self):
        print(" !!! TABLE: empty", self.categ, "table!")
        self.data = [["-"] * (5 if self.categ == "chem" else 2)]

    def read_data(self, ws):
        locator, extractor = self._get_parsers()
        result = locator(ws, self.flnm)
        if not result:
            self.empty()
            return
        if len(result) > 2:
            row, col, extrobj = result
            self.data = extractor(ws, extrobj, row, col, self.flnm)
        else:
            row, col = result
            self.data = extractor(ws, row, col, self.flnm)
        if not len(self.data):
            self.empty()

    def reparse_data(self):
        rp = []
        for line in self.data:
            if not line:
                continue
            if self.categ == "chem":
                units = line[3].split(" ")
                quant = units[0].replace(",", ".")
                try:
                    quant = str(float(quant)).replace(".", ",")
                except ValueError:
                    quant = line[3]
                    unit = ""
                else:
                    unit = " ".join(units[1:])
                repline = [line[4], line[0], unit, quant]
            else:
                cand, refname = extract_inventory_numbers(line[0])
                if len(line) > 1:
                    qnt = line[1]
                else:
                    qnt = ""
                repline = ["/".join(cand), refname.strip(), qnt]
                if self.categ == "inst":
                    repline.insert(2, "")
            rp.append(repline)
        if not rp:
            rp = [[""]*(4 if self.categ == "chem" else 3)]
        self.reparsed = np.array(rp)

    def _get_parsers(self):
        if self.categ == "chem":
            from SciProjects.xlcrawl.extract_chemicals import (
                locate_table_header as locator,
                extract_matrix as extractor
            )
        elif self.categ == "item":
            from SciProjects.xlcrawl.extract_items import (
                locate_table as locator,
                extract_matrix as extractor
            )
        else:
            from SciProjects.xlcrawl.extract_instrument import (
                locate_table as locator,
                extract_matrix as extractor
            )
        return locator, extractor

    def dump(self, xlwb: xl.Workbook):
        if self.reparsed is None:
            msg = " ".join(("Attempting to dump unreparsed", self.categ, "data"))
            raise RuntimeError(msg)
        ws = xlwb.get_sheet_by_name(self.categ)
        Xn, Yn = self.reparsed.shape
        cellz = np.array(list(ws.iter_rows(
            min_row=6, max_row=5 + Xn, min_col=1, max_col=Yn
        )))
        for c, d in zip(cellz.ravel(), self.reparsed.ravel()):
            d = str(d).replace("None", "").replace("-", "")
            c.value = int(d) if d.isdigit() else d


class Header(NamelessAbstraction):

    ranges = {
        "djnum": "4", "djname": "5", "owner": "6",
        "mname": "9", "akknums": "10", "mnum": "11",
        "mernok": "14", "mtasz": "15", "mhour": "16",
        "hmernok": "17", "hmtasz": "18",
        "techn": "19", "ttasz": "20", "thour": "21",
        "htechn": "22", "httasz": "23"
    }
    ranges = {k: "B"+v for k, v in ranges.items()}

    def __init__(self, data, path, flnm):
        super().__init__(path, flnm, "head", data)
        if isinstance(data, Worksheet):
            self.data = [None, None]
            self.data[0] = self.get_timereq_data(data, flnm)
            self.data[1] = self.get_dj_data(data, flnm)
            self.reparsed = self.data

    @staticmethod
    def get_timereq_data(ws, flnm):
        from SciProjects.xlcrawl.extract_timereq import extract_ws
        from SciProjects.xlcrawl.util import Allomany

        staffdb = Allomany()

        mtime, ttime, owner, tname = tuple(extract_ws(ws, flnm))
        mtasz, ttasz = staffdb.tasz(owner), staffdb.tasz(tname)
        mhely, thely = staffdb.helyettes(owner), staffdb.helyettes(tname)
        mhtasz, thtasz = staffdb.tasz(mhely), staffdb.tasz(thely)
        return ((owner, mtasz, mtime, mhely, mhtasz),
                (tname, ttasz, ttime, thely, thtasz))

    @staticmethod
    def get_dj_data(ws, flnm):
        from SciProjects.xlcrawl.util import DJ
        from SciProjects.xlcrawl.scrape_dj_name_and_number import (
            walk_columns, extract_djnum_easy, infer_djnum_from_string
        )
        djname, mname, akkr = walk_columns(ws, flnm)
        djnum = extract_djnum_easy(ws)
        djdb = DJ(projectroot + "Díjjegyzék.xlsx")
        if djnum is None:
            cands, rfnames, djnums, ps = infer_djnum_from_string(djname, flnm)
            for djnum in djnums:
                if str(djnum) in djname:
                    break
            else:
                if djnum:
                    djnum = djnums[0]
                    djname = rfnames[0]
                else:
                    djnum = ""
        else:
            djname = djdb.djnames[djnum] if djnum else ""
        djname = djname.replace(str(djnum), "")
        djname = "".join((djname[:8]
                          .replace(str(djnum), "")
                          .replace(".", "")
                          .strip(),
                          djname[8:]))
        djname = (djname
                  .replace("   ", " ")
                  .replace("  ", " ")
                  .strip())
        mnum = djdb.munumbers[djnum] if djnum else ""
        mname = djname
        return [djnum, djname, mnum, mname]

    def dump(self, xlwb: xl.Workbook):

        def cell(rng, val):
            if isinstance(val, str):
                val = "" if val.lower() in ("none", "-") else val
                val = int(val) if val.isdigit() else val
            if val is None:
                val = ""
            ws[self.ranges[rng]].value = val

        ws = xlwb.get_sheet_by_name("head")
        owner, mtasz, mtime, mhely, mhtasz = self.data[0][0]
        tname, ttasz, ttime, thely, thtasz = self.data[0][1]
        djnum, djname, mnum, mname = self.data[1]

        cell("djnum", djnum)
        cell("djname", djname)
        cell("mnum", mnum)
        cell("mname", djname)
        cell("mernok", owner)
        cell("mtasz", mtasz)
        cell("mhour", mtime)
        cell("hmernok", mhely)
        cell("hmtasz", mhtasz)
        cell("techn", tname)
        cell("ttasz", ttasz)
        cell("thour", ttime)
        cell("htechn", thely)
        cell("httasz", thtasz)
        cell("owner", owner)


class Method:

    def __init__(self, flpath, ws=None):
        path, flnm = _slicepath(flpath)
        self.path = path
        self.flnm = flnm
        self.outwb = None
        if not ws:
            xlin = xl.load_workbook(flpath)
            self.ws = xlin.worksheets[0]
        else:
            self.ws = ws
        self.sanity_check()
        self.head = Header(self.ws, path, flnm)
        self.chem = Table(self.ws, path, flnm, "chem")
        self.inst = Table(self.ws, path, flnm, "inst")
        self.item = Table(self.ws, path, flnm, "item")

    @staticmethod
    def load(flnm):
        with gzip.open(pickleroot + flnm) as handle:
            return pickle.load(handle)

    def sanity_check(self):
        return "Melléklet" in (self.ws["A1"].value, self.ws["A2"].value)

    def dump(self, newroot):
        outwb = xl.load_workbook(templateroot + "template.xlsx")
        self.save()
        os.chdir(newroot)
        self.head.dump(outwb)
        self.chem.dump(outwb)
        self.item.dump(outwb)
        self.inst.dump(outwb)
        try:
            os.makedirs(f"{self.path}")
        except FileExistsError:
            pass
        outwb.save(f"{self.path}/{self.flnm}")

    def save(self, path=None):
        if path is None:
            path = pickleroot
        with gzip.open(path + self.flnm[:-4]+"pkl.gz", "wb") as handle:
            pickle.dump(self, handle)


def cleanup(dstroot, force=False):
    import shutil
    msg = "\n".join(("You are attempting to empty the directory:",
                     dstroot,
                     "Are you sure? [yes] > "))
    if not force:
        if input(msg) != "yes":
            return
    for entry in os.listdir(dstroot):
        if os.path.isfile(entry):
            os.remove(dstroot + "/" + entry)
        else:
            shutil.rmtree(dstroot + "/" + entry)


def crawl(destination, source=sourceroot):
    cleanup(destination)

    for xlworksheet, xlpath in iter_flz(source):
        method = Method(xlpath, xlworksheet)
        method.dump(destination)

if __name__ == '__main__':
    start = time.time()
    destroot = projectroot + "reparsed/"
    crawl(destination=destroot, source=sourceroot)
    print(f"--- FINITE --- ({time.time()-start:>.4f})")
