import os
import pickle
from collections import defaultdict

import openpyxl as xl
from SciProjects.xlcrawl import projectroot
from SciProjects.xlcrawl.util import Allomany, DJ, headranges

from openpyxl.worksheet import Worksheet

personell = Allomany()
dj = DJ()
times = pickle.load(open(projectroot + "names.pkl", "rb"))
times[101] = [10, 105]
times[196] = [0, 0]

vsum = open(projectroot + "Vizsg2016ossz.csv").read().split("\n")
vsum = {k: v for k, v in (map(int, l.split("\t")) for l in vsum if l)}


def getwb(path):
    wb = xl.load_workbook(path)
    if "head" not in wb.sheetnames:
        print("head not in wb.sheetnames")
        return
    testme = wb.get_sheet_by_name("head")["A1"].value[:9]
    if "Önköltség" != testme:
        print(testme, "!= Önköltség")
        return
    return wb


def inject_hpeople(xlwb: xl.Workbook, flnm):

    def correct(name):
        if name is None:
            return "-"
        name = name.strip()
        if name in ("-", ""):
            return "-"
        while name not in personell.data:
            if "Marusnikné" in name:
                return "Marusnikné Sárközi Ágnes"
            elif "Szabó Ferenc" in name:
                return "Szabó Ferenc"
            name = input("FURA @ {}\nKi ez? [{}] > ".format(flnm, name))
        return name

    def parse_helyettes(name):
        chainz = []
        h = personell.helyettes(name)
        hs = h.split("; ")
        for nm in hs:
            nm = correct(nm)
            tasz = personell.tasz(nm)
            chainz.append(f"{nm} ({tasz})")
        return ", ".join(chainz)

    head = xlwb.get_sheet_by_name("head")
    mernok = correct(head[headranges["mernok"]].value)
    technikus = correct(head[headranges["techni"]].value)
    djn = str(head[headranges["djn"]].value).replace(".", "")
    head[headranges["mernok"]].value = mernok
    head[headranges["mtasz"]].value = personell.tasz(mernok)
    head[headranges["techni"]].value = technikus
    head[headranges["ttasz"]].value = personell.tasz(technikus)
    head[headranges["hmernok"]].value = parse_helyettes(mernok)
    head[headranges["hmtasz"]].value = ""
    head[headranges["htechni"]].value = parse_helyettes(technikus)
    head[headranges["httasz"]].value = ""
    mtime = str(head[headranges["mtime"]].value)
    ttime = str(head[headranges["ttime"]].value)
    if not djn.isdigit():
        return
    mn = dj.munumbers[int(djn)]
    if not mtime.isdigit() and djn.isdigit():
        mtime = times.get(mn, None)
        if mtime is None:
            print("MISSING MTIME IN", flnm, "DJZ:", dj.mu_to_dj(mn))
        else:
            mtime = mtime[0]
    if not ttime.isdigit() and djn.isdigit():
        ttime = times.get(mn, None)
        if ttime is None:
            print("MISSING TTIME IN", flnm, "DJZ:", dj.mu_to_dj(mn))
        else:
            ttime = ttime[1]
    head[headranges["mtime"]] = mtime
    head[headranges["ttime"]] = ttime


def gettimes(ws):
    djn = ws[headranges["djn"]].value
    if not str(djn).isdigit():
        return None
    mnum = dj.munumbers[int(djn)]
    mtime, ttime = ws[headranges["mtime"]].value, ws[headranges["ttime"]].value
    if mtime is None or ttime is None:
        return None
    return {mnum: [mtime, ttime]}


def printem(headws, flnm):
    chainz = []
    djnum = headws["B4"].value
    djname = headws["B5"].value
    owner = headws["B6"].value
    chainz.append(str(djnum).lower()
                  .replace(" ", "")
                  .replace("none", "")
                  .replace("nincs", "")
                  .replace("-", ""))
    chainz.append(djname)
    chainz.append(owner)
    chainz.append(flnm)
    out = "\t".join(chainz) + "\n"
    # print(out)
    return out


def getduper(headws, flnm):
    djnum = str(headws["B4"].value).strip()
    djname = headws["B5"].value
    owner = headws["B6"].value
    if "," in djnum:
        print("-"*80)
        print(f"{owner}: {djnum}-{djname} @ {flnm}")


def fill_owner(ws, flnm):
    v = str(ws[headranges["owner"]])
    if len(str(ws[headranges["mernok"]].value)) < 7:
        print("!! UNFIXABLE:", flnm)
    if len(v) < 7:
        print("FIXING", flnm)
    ws[headranges["owner"]].value = ws[headranges["mernok"]].value


def fill_vizsgsum(ws: Worksheet, flnm):
    ws.protection.disable()
    djn = str(ws[headranges["djn"]].value).replace(".", "")
    if not djn.isdigit():
        print("SKIPPED", flnm)
        return
    insert = "2016-ban ezt a vizsgálatot {} alkalommal végezték."
    ws[headranges["vizsgsum"]].value = insert.format(vsum[int(djn)])


def move_to_destination(xlstream, root):
    unknown = 1
    djns = defaultdict(int)

    for flnm in xlstream:
        wb = getwb(root + flnm)
        ws = wb["head"]
        djn = str(ws[headranges["djn"]].value).strip().replace(".", "")
        djname = ws[headranges["djname"]].value
        if not djn.isdigit():
            djn = "U{:0>2}".format(unknown)
            unknown += 1
        ndjn = "{:3>0}".format(djn)
        djns[djn] += 1
        if djns[djn] > 1:
            djn += "{:0>2}-".format(djns[djn])
        newpath = projectroot + "OUTPUT/" + ndjn + "-" + djname.replace("/", "-") + ".xlsx"
        wb.save(newpath)


def main():
    root = projectroot + "FINAL/"
    xlstream = [fl for fl in os.listdir(root) if fl[0] != "~" and fl[-5:] == ".xlsx"]
    allflz = len(xlstream)
    strln = len(str(allflz))
    for i, flnm in enumerate(xlstream, start=1):
        # print(f"\rDoing {i:>{strln}}/{allflz} - {flnm}", end="")
        wb = getwb(root + flnm)
        if wb is None:
            print("INVALID FILE:", flnm)
            continue
        # printem(wb.get_sheet_by_name("head"), flnm)
        fill_vizsgsum(wb["head"], flnm)
        wb.save(root + flnm)


if __name__ == '__main__':
    move_to_destination((fl for fl in os.listdir(projectroot + "FINAL/")
                         if fl[0] != "~" and fl[-5:] == ".xlsx"),
                        projectroot + "FINAL/")
