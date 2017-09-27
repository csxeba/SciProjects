import openpyxl as xl

from SciProjects.fruits import projectroot


def parse_xlfl(path):
    wb = xl.load_workbook(path, data_only=True)
    for sheetname in wb.sheetnames:
        ws = wb.get_sheet_by_name(sheetname)
        ID = sheetname
        for nID in sheetname.split(" "):
            if nID.isdigit() and len(nID) == 8:
                # ID = nID
                break
        else:
            print(ID, end="\t")
            for rown in range(23, 33):
                print(ws[f"R{rown}"].value, end="\t")
            print()


if __name__ == '__main__':
    parse_xlfl(projectroot + "Kozma_nyers/Pálinka.adatbázis.2012_2013.xlsx")
